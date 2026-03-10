from pathlib import Path
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook

INDIR_CD_INFO = Path("source/raw/post1790_cd")
INDIR_RAW = Path("source/raw/post1790_cd/orig")
INDIR_PRE1790 = Path("output/derived/prescrape/pre1790")
OUTDIR = Path("output/analysis/pre1790/debt_per_year")
OUTDIR.mkdir(parents=True, exist_ok=True)
(OUTDIR / "year_month").mkdir(parents=True, exist_ok=True)


def Main():
    PlotPre1790ByYear()
    PlotPre1790PctByYear()

    cd_info_path = INDIR_CD_INFO / "cd_info.csv"
    if not cd_info_path.exists():
        print(f"Skipping per-state plots: {cd_info_path} not found (Task 4 item)")
        return
    cd_info = pd.read_csv(cd_info_path)
    all_states_frames = []

    for _, row in cd_info.iterrows():
        state_df = LoadStateXlsx(row)
        if state_df is None or state_df.empty:
            continue
        state_df = CleanDf(state_df)
        missing_sum = HandleMissingInfo(state_df)
        state_df = state_df.dropna(subset=["year", "month"])
        state_df["year"] = state_df["year"].astype(str)
        state_df["month"] = state_df["month"].astype(str).str.zfill(2)
        state_df["year_month"] = state_df["year"] + "-" + state_df["month"]
        gdf = GroupByYearMonth(state_df)
        gdf = GroupPost1795(gdf, missing_sum)
        PlotDebt(gdf, row["state"])
        all_states_frames.append(gdf[["year_month", "total_amt"]].copy())

    if all_states_frames:
        all_states = pd.concat(all_states_frames, ignore_index=True)
        us_gdf = all_states.groupby("year_month", as_index=False)["total_amt"].sum()
        us_gdf = GroupPost1795(us_gdf, 0)
        PlotDebt(us_gdf, "US")


def LoadStateXlsx(row):
    file_path = Path(row["file_path"])
    if not file_path.exists():
        file_path = INDIR_RAW / file_path.name
    if not file_path.exists():
        return None
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb["Sheet1"]
        first_row, last_row = int(row["first_row"]), int(row["last_row"])
        n = last_row - first_row + 1

        year = MergeCols(ws, str(row["year_col"]), first_row, last_row)
        month = MergeCols(ws, str(row["month_col"]), first_row, last_row)
        dollars = MergeCols(ws, str(row["dollars_col"]), first_row, last_row)
        cents_col = row.get("cents_col")
        if pd.notna(cents_col):
            cents_raw = MergeCols(ws, str(cents_col), first_row, last_row)
            cents = [float(v) / 100 if v is not None else 0.0 for v in cents_raw]
        else:
            cents = [0.0] * n
        wb.close()

        return pd.DataFrame({"year": year, "month": month, "dollars": dollars, "cents": cents})
    except Exception:
        return None


def MergeCols(ws, col_str, first_row, last_row):
    n = last_row - first_row + 1
    result = [None] * n
    for col_letter in col_str.split(","):
        col_letter = col_letter.strip()
        for i, r in enumerate(range(first_row, last_row + 1)):
            val = ws[f"{col_letter}{r}"].value
            if result[i] is None and val is not None:
                result[i] = val
    return result


def CleanDf(df):
    for col in ["dollars", "cents"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df


def HandleMissingInfo(df):
    missing = df[df["year"].isna() | df["month"].isna()]
    return (missing["dollars"] + missing["cents"]).sum()


def GroupByYearMonth(df):
    df["total_amt"] = df["dollars"] + df["cents"]
    return df.groupby("year_month", as_index=False)["total_amt"].sum()


def GroupPost1795(gdf, missing_sum):
    gdf["_sort_key"] = pd.to_datetime(gdf["year_month"], format="%Y-%m", errors="coerce")
    post1795 = gdf[gdf["_sort_key"] > pd.Timestamp("1795-01-01")]
    pre1795 = gdf[gdf["_sort_key"] <= pd.Timestamp("1795-01-01")].drop(columns="_sort_key")
    post1795_sum = post1795["total_amt"].sum()
    extra_rows = []
    if post1795_sum > 0:
        extra_rows.append({"year_month": "post-1795", "total_amt": post1795_sum})
    if missing_sum > 0:
        extra_rows.append({"year_month": "missing info", "total_amt": missing_sum})
    if extra_rows:
        pre1795 = pd.concat([pre1795, pd.DataFrame(extra_rows)], ignore_index=True)
    return pre1795


def PlotDebt(gdf, state):
    fig, ax = plt.subplots(figsize=(14, 6))
    ax.bar(range(len(gdf)), gdf["total_amt"], tick_label=gdf["year_month"].tolist())
    ax.set_xlabel("Year/Month")
    ax.set_ylabel("Total Debt (in dollars)")
    ax.set_title(f"{state.upper()} Debt Redeemed Per Year/Month")
    plt.xticks(rotation=90, fontsize=7)
    plt.tight_layout()
    out = OUTDIR / "year_month" / f"{state.upper()}_debt_redeemed_per_year_month.png"
    plt.savefig(out, dpi=150)
    plt.close()


def PlotPre1790ByYear():
    pre1790 = pd.read_csv(INDIR_PRE1790 / "pre1790_cleaned.csv")
    pre1790["year"] = pre1790.get("date of the certificate | year")
    pre1790["dollars"] = pd.to_numeric(pre1790.get("amount | dollars"), errors="coerce").fillna(0)
    cents_raw = pre1790.get("amount | 90th", pd.Series(dtype=str))
    cents_raw = cents_raw.astype(str).str.split(".").str[0].str.replace("/", "")
    pre1790["cents"] = pd.to_numeric(cents_raw, errors="coerce").fillna(0) / 100
    pre1790.loc[pre1790["cents"] >= 100, "cents"] = 0
    pre1790["total_amt"] = pre1790["dollars"] + pre1790["cents"]

    gdf = pre1790.groupby("year", as_index=False)["total_amt"].sum()
    gdf["year"] = gdf["year"].fillna(0).astype(int).astype(str)
    gdf.loc[gdf["year"] == "0", "year"] = "no year"
    gdf = gdf.sort_values("year")
    gdf["total_amt_m"] = (gdf["total_amt"] / 1e6).round(2)

    fig, ax = plt.subplots(figsize=(12, 6))
    bars = ax.bar(range(len(gdf)), gdf["total_amt_m"], tick_label=gdf["year"].tolist())
    ax.bar_label(bars, labels=gdf["total_amt_m"].tolist(), fontsize=8, padding=2)
    ax.set_ylim(0, gdf["total_amt_m"].max() * 1.15)
    ax.set_xlabel("Year")
    ax.set_ylabel("Amount (in million dollars)")
    ax.set_title("Debt Certificate Total Per Year")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    plt.savefig(OUTDIR / "pre1790_debt_certificate_amts_per_year.png", dpi=150)
    plt.close()


def PlotPre1790PctByYear():
    pre1790 = pd.read_csv(INDIR_PRE1790 / "pre1790_cleaned.csv")
    pre1790["year"] = pre1790.get("date of the certificate | year")
    pre1790["dollars"] = pd.to_numeric(pre1790.get("amount | dollars"), errors="coerce").fillna(0)
    cents_raw = pre1790.get("amount | 90th", pd.Series(dtype=str))
    cents_raw = cents_raw.astype(str).str.split(".").str[0].str.replace("/", "")
    pre1790["cents"] = pd.to_numeric(cents_raw, errors="coerce").fillna(0) / 100
    pre1790.loc[pre1790["cents"] >= 100, "cents"] = 0
    pre1790["total_amt"] = pre1790["dollars"] + pre1790["cents"]

    gdf = pre1790.groupby("year", as_index=False)["total_amt"].sum()
    gdf["year"] = gdf["year"].fillna(0).astype(int).astype(str)
    gdf.loc[gdf["year"] == "0", "year"] = "no year"
    gdf = gdf.sort_values("year")
    total = gdf["total_amt"].sum()
    gdf["percent"] = (gdf["total_amt"] / total).round(2)

    fig, ax = plt.subplots(figsize=(12, 6))
    bars = ax.bar(range(len(gdf)), gdf["percent"], tick_label=gdf["year"].tolist())
    ax.bar_label(bars, labels=gdf["percent"].tolist(), fontsize=8, padding=2)
    ax.set_ylim(0, gdf["percent"].max() * 1.15)
    ax.set_xlabel("Year")
    ax.set_ylabel("Amount (Percentage of Total)")
    ax.set_title("Percent of Debt Certificate Total Per Year")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    plt.savefig(OUTDIR / "pre1790_debt_certificate_percent_per_year.png", dpi=150)
    plt.close()


if __name__ == "__main__":
    Main()
